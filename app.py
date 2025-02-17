from flask import Flask, request, jsonify, render_template
import openpyxl
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage
import os

app = Flask(__name__, static_folder="static", template_folder="static")

# Nome do arquivo da planilha
ARQUIVO_RESERVAS = "reservas_tablets.xlsx"

# Função para verificar conflitos de horário
def verificar_conflito(data, hora_inicio, hora_fim):
    try:
        wb = openpyxl.load_workbook(ARQUIVO_RESERVAS)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            if row[2] == data and row[3] == hora_inicio and row[4] == hora_fim:
                return True  # Conflito encontrado

        return False  # Nenhum conflito
    except FileNotFoundError:
        return False  # Se o arquivo não existir, não há conflito

# Função para salvar os dados na planilha
def salvar_dados(nome, quantidade, data, hora_inicio, hora_fim):
    if verificar_conflito(data, hora_inicio, hora_fim):
        return False  # Bloqueia reserva duplicada no mesmo horário e data

    try:
        wb = openpyxl.load_workbook(ARQUIVO_RESERVAS)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Quantidade de Tablets", "Data", "Hora Início", "Hora Fim"])

    ws.append([nome, quantidade, data, hora_inicio, hora_fim])
    wb.save(ARQUIVO_RESERVAS)

    enviar_email(ARQUIVO_RESERVAS)
    return True

# Função para enviar o e-mail com a planilha
def enviar_email(arquivo):
    email_remetente = os.getenv("EMAIL_USER", "cef102ti@gmail.com")
    senha = os.getenv("EMAIL_PASS", "ekmx nhnc ufxw zfvj")
    email_destino = "cef102ti@gmail.com"

    msg = EmailMessage()
    msg["Subject"] = "Reserva de Tablets"
    msg["From"] = email_remetente
    msg["To"] = email_destino
    msg.set_content("Segue em anexo a planilha com a reserva de tablets.")

    with open(arquivo, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=arquivo,
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(email_remetente, senha)
        server.send_message(msg)
  
    print("E-mail enviado com sucesso!")

# Rota da página principal
@app.route("/")
def index():
    return render_template("index.html")

# Rota para processar reservas
@app.route('/reservar', methods=['POST'])
def reservar():
    data = request.json
    nome = data.get("nome")
    quantidade = data.get("quantidade")
    data_reserva = data.get("data")
    hora_inicio = data.get("hora_inicio")
    hora_fim = data.get("hora_fim")

    if not all([nome, quantidade, data_reserva, hora_inicio, hora_fim]):
        return jsonify({"error": "Preencha todos os campos!"}), 400

    if salvar_dados(nome, quantidade, data_reserva, hora_inicio, hora_fim):
        return jsonify({"message": "Reserva salva e e-mail enviado com sucesso!"})
    else:
        return jsonify({"error": "Já existe uma reserva nesse horário!"}), 400

if __name__ == "__main__":
    app.run(debug=True)
